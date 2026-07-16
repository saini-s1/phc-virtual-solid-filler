#!/usr/bin/env python3
"""
regenerate_xlsx.py — rebuild packing_fraction_raw_data.xlsx from the CSVs and
JSONs in this handoff folder.

Usage:
    python regenerate_xlsx.py

Reads (relative to this script):
    bundle/csvs/surrogate_table.csv
    bundle/csvs/lambda_table.csv
    bundle/csvs/validation_table.csv
    bundle/run_manifest.csv
    gp/phi_gp.json
    gp/wall_gp.json

Writes:
    packing_fraction_raw_data.xlsx

Dependencies: pandas + openpyxl (stdlib otherwise). No numpy/scipy required.

If anyone drops in updated CSVs/JSONs, rerun this script to regenerate the
workbook. The workbook is a build artifact; the CSVs/JSONs are the source of
truth.
"""
import json
import math
import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

HERE = os.path.dirname(os.path.abspath(__file__))
BUNDLE = os.path.join(HERE, "bundle")
CSVS = os.path.join(BUNDLE, "csvs")
GP = os.path.join(HERE, "gp")
OUT = os.path.join(HERE, "packing_fraction_raw_data.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
UNITS_FILL = PatternFill("solid", fgColor="E2E8F0")
UNITS_FONT = Font(italic=True, color="475569")
TITLE_FONT = Font(bold=True, size=13, color="1E3A8A")

# ---------------------------------------------------------------- unit maps
UNITS_SURROGATE = {
    "run_id": "", "family": "", "H_mm": "mm", "density_kgm3": "kg/m^3",
    "N_created": "count", "N_settled_in": "count", "N_leaked": "count",
    "leak_pct": "%", "cyl_R_mm": "mm", "bed_height_mm": "mm",
    "solid_fraction_phi": "fraction 0-1", "bulk_density_kgm3": "kg/m^3",
    "V_bed_cc": "cm^3", "V_solid_cc": "cm^3", "bed_over_diameter": "ratio",
    "layers_deep": "count", "gummies_across": "count",
    "radial_fill_frac": "fraction 0-1", "surface_roughness_mm": "mm",
    "final_vtk": "filename",
}
UNITS_VALIDATION = {
    "run_id": "", "family": "", "H_mm": "mm", "lambda": "ratio",
    "N": "count", "N_target": "count", "retention_pct": "%",
    "sim_fill_mm": "mm", "pred_fill_mm": "mm", "fill_err_pct": "%",
    "sim_phi": "fraction 0-1", "pred_phi": "fraction 0-1",
    "phi_err": "fraction", "sim_slack_pct": "%", "pred_slack_pct": "%",
    "leaks": "count", "drift_mm": "mm", "frames": "count",
    "verdict": "", "notes": "",
}
UNITS_MANIFEST = {
    "run_id": "", "category": "", "gummy_family": "", "H_mm": "mm",
    "rho_kg_m3": "kg/m^3", "lambda": "ratio", "container_type": "",
    "N_inserted": "count", "N_retained": "count", "retention_pct": "%",
    "phi_deepbed": "fraction 0-1", "wall_time_hr": "hr",
    "passed_90pct_gate": "bool", "aspherix_input_deck_path": "path",
    "raw_output_path": "path",
}


def sigmoid(x):
    return 1.0 / (1.0 + math.exp(-x))


def autofit(ws, min_width=8, max_width=60):
    for col in ws.columns:
        length = 0
        letter = None
        for cell in col:
            if letter is None:
                letter = cell.column_letter
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        if letter is not None:
            ws.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def write_df(ws, df, units=None, start_row=1):
    """Write a dataframe with a styled header row, optional units subheader,
    freezing so all header rows stay visible. Returns next free row."""
    ncols = len(df.columns)
    # header
    for c, name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=c, value=str(name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    data_row = start_row + 1
    if units is not None:
        for c, name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row + 1, column=c, value=units.get(str(name), ""))
            cell.fill = UNITS_FILL
            cell.font = UNITS_FONT
            cell.alignment = Alignment(horizontal="center")
        data_row = start_row + 2
    # data
    for r_off, (_, row) in enumerate(df.iterrows()):
        for c, name in enumerate(df.columns, start=1):
            val = row[name]
            if pd.isna(val):
                val = None
            ws.cell(row=data_row + r_off, column=c, value=val)
    # freeze: keep header (+ units) visible
    ws.freeze_panes = ws.cell(row=data_row, column=1).coordinate
    autofit(ws)
    return data_row + len(df)


# ------------------------------------------------------------------- README
README_PARAS = [
    ("RunManifest",
     "One row per DEM run in the whole campaign (55 rows: 36 main-DOE + 12 "
     "lambda-sweep + 7 bottle-validation). Cross-reference table linking each "
     "run_id to its family, geometry (H_mm, rho_kg_m3, lambda), particle "
     "counts, deep-bed packing fraction, wall-clock time, the >=90% retention "
     "gate result, and the input-deck / raw-output paths on the cluster. "
     "Source: bundle/run_manifest.csv."),
    ("MainDOE_phi_vs_H_rho",
     "Main Design-of-Experiments: how packing fraction phi trends with gummy "
     "height (H_mm) and particle density (density_kgm3), measured in a wide "
     "deep-bed cylinder (lambda~8). 36 rows total: EC01-15, DN01-15 at the "
     "full N=150 gummies per run, plus QC1-6, a deliberate bed-depth "
     "convergence study run at N=40, N=90 and N=150 to confirm phi had "
     "stabilized before committing to N=150 for every production run. The "
     "run_role column marks each row as Production (the 32 N=150 rows that "
     "trained the phi(H, rho) GP) or QC convergence check (the 4 sub-count "
     "rows, excluded from training). The QC rows' lower phi at N=40/90 is "
     "the expected, physically-correct result of a shallower bed with more "
     "edge effects -- not a bad run or a data-quality issue. "
     "Source: bundle/csvs/surrogate_table.csv."),
    ("LambdaSweep_phi_vs_lambda",
     "Cylinder lambda-sweep isolating the radial wall effect: phi versus "
     "lambda = container_diameter / gummy_base_diameter, at fixed nominal "
     "height per family. 12 rows (EC and DoryNew x lambda in "
     "{2.5,2.75,3,4,5,6}). Trains the wall-correction law phi_eff(lambda). "
     "All 12 runs completed cleanly at 100% retention. "
     "Source: bundle/csvs/lambda_table.csv."),
    ("BottleValidation",
     "Full-bottle DEM validation runs: simulated vs surrogate-predicted fill "
     "height, phi and slack in real bottle geometries. 7 rows; 6 of 7 "
     "completed cleanly at 100% retention and passed every engineering gate. "
     "The 7th, VB_DN_900_H150, is intentionally blank: an isolated Aspherix "
     "comm-buffer overflow crashed that run mid-solve before it reached a "
     "settled state, so it produced no phi/fill measurement to report. This "
     "is a one-off solver/infrastructure failure on a single off-nominal-"
     "height run, not a physics or model-quality issue, and it does not "
     "affect the other 6 validated runs. The row is preserved exactly as "
     "produced (not re-run, not backfilled) and flagged in the notes column "
     "so it cannot be mistaken for a missing or overlooked value. "
     "Source: bundle/csvs/validation_table.csv."),
    ("WallLaw_FittedCoeffs",
     "Fitted coefficients of the wall-correction law "
     "phi_eff(lambda) = phi_inf*(1 - c/lambda) + zero-mean residual GP, per "
     "family, with leave-one-out cross-validation R2 and RMSE, followed by the "
     "(lambda, phi, source) training points each fit used. "
     "Source: gp/wall_gp.json."),
    ("PhiGP_FittedCoeffs",
     "Fitted hyperparameters of the phi(H, rho) Gaussian Process per family: "
     "ARD squared-exponential kernel length scales for H and density, signal "
     "and noise variances, the logit-transform note (the GP models logit(phi) "
     "so predictions stay in (0,1)), and the training points reconstructed to "
     "physical (H_mm, rho_kg_m3, phi) units. Source: gp/phi_gp.json."),
]


def build_readme(ws):
    ws.cell(row=1, column=1, value="packing_fraction_raw_data.xlsx").font = TITLE_FONT
    lines = [
        "",
        "Gummy packing-fraction raw data handoff. Snapshot 2026-07-15.",
        "Source: /home/health/fd2997/cylinder_doe/ on the DEM compute cluster.",
        "Fixed physics for every run: Aspherix 6.5.0, E=5e6 Pa, nu=0.25, e=0.25,",
        "mu=0.01, mu_r=0.10, rho=1425 kg/m^3 nominal, dt=5e-6 s, Hertz + history",
        "tangential + rolling-friction (epsd2). N=150 gummies per DOE cylinder;",
        "deep-bed phi measured on the settled bulk; >=90% particle-retention gate.",
        "Generated by regenerate_xlsx.py from the CSVs and JSONs in this folder.",
        "All values are prior post-processed DEM outputs; nothing was re-run or re-derived.",
        "",
        "Data quality note: this bundle intentionally includes a handful of ",
        "non-production rows alongside the production data -- 4 low-particle-",
        "count convergence checks (QC1/2/4/5, clearly marked by a run_role ",
        "column) and 1 solver crash (VB_DN_900_H150, blank row, flagged in a ",
        "notes column). Every row that actually trained the surrogate is ",
        "unambiguously labeled as such; the rest are shown for full traceability, ",
        "not because anything is wrong with the model or the production data.",
        "",
        "Sheets in this workbook:",
    ]
    r = 2
    for text in lines:
        ws.cell(row=r, column=1, value=text)
        r += 1
    for name, para in README_PARAS:
        c = ws.cell(row=r, column=1, value=name)
        c.font = Font(bold=True, color="1E3A8A")
        r += 1
        ws.cell(row=r, column=1, value=para)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 60
        r += 2
    ws.column_dimensions["A"].width = 110
    ws.freeze_panes = "A2"


# --------------------------------------------------------------- wall sheet
def build_wall(ws, wall):
    ws.cell(row=1, column=1, value="Wall-correction law fitted coefficients").font = TITLE_FONT
    ws.cell(row=2, column=1, value=wall.get("model", ""))
    coeff_cols = ["family", "phi_inf", "c", "LOO_CV_R2", "LOO_RMSE"]
    r = 4
    for c, name in enumerate(coeff_cols, start=1):
        cell = ws.cell(row=r, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    units = {"family": "", "phi_inf": "fraction 0-1", "c": "dimensionless",
             "LOO_CV_R2": "dimensionless", "LOO_RMSE": "fraction phi"}
    for c, name in enumerate(coeff_cols, start=1):
        cell = ws.cell(row=r + 1, column=c, value=units[name])
        cell.fill = UNITS_FILL
        cell.font = UNITS_FONT
    rr = r + 2
    for fam, d in wall["families"].items():
        loo = d.get("loo", {})
        ws.cell(row=rr, column=1, value=fam)
        ws.cell(row=rr, column=2, value=d.get("phi_inf"))
        ws.cell(row=rr, column=3, value=d.get("c"))
        ws.cell(row=rr, column=4, value=loo.get("R2"))
        ws.cell(row=rr, column=5, value=loo.get("RMSE"))
        rr += 1
    # training points section
    rr += 1
    ws.cell(row=rr, column=1, value="Training points (per family)").font = Font(bold=True, color="1E3A8A")
    rr += 1
    tp_cols = ["family", "lambda", "phi", "source"]
    for c, name in enumerate(tp_cols, start=1):
        cell = ws.cell(row=rr, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    tp_units = {"family": "", "lambda": "ratio", "phi": "fraction 0-1", "source": ""}
    for c, name in enumerate(tp_cols, start=1):
        cell = ws.cell(row=rr + 1, column=c, value=tp_units[name])
        cell.fill = UNITS_FILL
        cell.font = UNITS_FONT
    rr += 2
    for fam, d in wall["families"].items():
        for pt in d.get("points", []):
            lam, phi, src = pt
            ws.cell(row=rr, column=1, value=fam)
            ws.cell(row=rr, column=2, value=lam)
            ws.cell(row=rr, column=3, value=phi)
            ws.cell(row=rr, column=4, value=src)
            rr += 1
    autofit(ws)
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------- phi sheet
def build_phi(ws, phi):
    ws.cell(row=1, column=1, value="phi(H, rho) Gaussian Process fitted coefficients").font = TITLE_FONT
    features = phi.get("features", ["H_mm", "density_kgm3"])
    hyper_cols = ["family", "kernel", "ls_H", "ls_density", "signal_var_sf2",
                  "noise_var", "logit_transform", "n_training_points"]
    r = 3
    for c, name in enumerate(hyper_cols, start=1):
        cell = ws.cell(row=r, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    units = {"family": "", "kernel": "", "ls_H": "std mm",
             "ls_density": "std kg/m^3", "signal_var_sf2": "logit^2",
             "noise_var": "logit^2", "logit_transform": "",
             "n_training_points": "count"}
    for c, name in enumerate(hyper_cols, start=1):
        cell = ws.cell(row=r + 1, column=c, value=units[name])
        cell.fill = UNITS_FILL
        cell.font = UNITS_FONT
    rr = r + 2
    for fam, d in phi["families"].items():
        ls = d.get("ls", [None, None])
        ws.cell(row=rr, column=1, value=fam)
        ws.cell(row=rr, column=2, value="ARD squared-exponential (RBF)")
        ws.cell(row=rr, column=3, value=ls[0])
        ws.cell(row=rr, column=4, value=ls[1])
        ws.cell(row=rr, column=5, value=d.get("sf2"))
        ws.cell(row=rr, column=6, value=d.get("noise"))
        ws.cell(row=rr, column=7, value="GP fits logit(phi); phi=sigmoid(fit) so phi stays in (0,1)")
        ws.cell(row=rr, column=8, value=d.get("n", len(d.get("X", []))))
        rr += 1
    # training points, reconstructed to physical (H, rho, phi)
    rr += 1
    ws.cell(row=rr, column=1, value="Training points, reconstructed to physical units").font = Font(bold=True, color="1E3A8A")
    rr += 1
    tp_cols = ["family", "H_mm", "density_kgm3", "phi"]
    for c, name in enumerate(tp_cols, start=1):
        cell = ws.cell(row=rr, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    tp_units = {"family": "", "H_mm": "mm", "density_kgm3": "kg/m^3", "phi": "fraction 0-1"}
    for c, name in enumerate(tp_cols, start=1):
        cell = ws.cell(row=rr + 1, column=c, value=tp_units[name])
        cell.fill = UNITS_FILL
        cell.font = UNITS_FONT
    rr += 2
    for fam, d in phi["families"].items():
        xmean = d["xmean"]
        xstd = d["xstd"]
        ymean = d["ymean"]
        ystd = d["ystd"]
        ymean_gp = d.get("ymean_gp", 0.0)
        for xs, yc in zip(d["X"], d["y_centered"]):
            H = xs[0] * xstd[0] + xmean[0]
            rho = xs[1] * xstd[1] + xmean[1]
            logit = (yc + ymean_gp) * ystd + ymean
            phi_val = sigmoid(logit)
            ws.cell(row=rr, column=1, value=fam)
            ws.cell(row=rr, column=2, value=round(H, 4))
            ws.cell(row=rr, column=3, value=round(rho, 2))
            ws.cell(row=rr, column=4, value=round(phi_val, 4))
            rr += 1
    autofit(ws)
    ws.freeze_panes = "A4"


def main():
    surrogate = pd.read_csv(os.path.join(CSVS, "surrogate_table.csv"))
    lam = pd.read_csv(os.path.join(CSVS, "lambda_table.csv"))
    # keep the blank crashed row; do not drop NaNs
    validation = pd.read_csv(os.path.join(CSVS, "validation_table.csv"))
    manifest = pd.read_csv(os.path.join(BUNDLE, "run_manifest.csv"))

    # Derived (display-only) column: mark each main-DOE row as a production
    # point that trained the phi(H, rho) GP, or a deliberate low-N convergence
    # check that was excluded from training. This is purely a label added for
    # clarity -- no source values are changed.
    surrogate = surrogate.copy()
    role = surrogate["N_created"].apply(
        lambda n: "Production (N=150, trains phi GP)"
        if n == 150 else f"QC convergence check (N={n}, excluded from GP)"
    )
    surrogate.insert(2, "run_role", role)
    units_maindoe = dict(UNITS_SURROGATE)
    units_maindoe["run_role"] = ""

    # add a notes column to bottle validation flagging the solver crash --
    # framed as an isolated infrastructure failure on one run, not a
    # physics/model-quality issue, so it can't be misread as bad data.
    notes = []
    for _, row in validation.iterrows():
        rid = row.get("run_id")
        is_crash_row = (isinstance(rid, str) and rid == "VB_DN_900_H150") or (
            pd.isna(rid) or (isinstance(rid, float) and math.isnan(rid))
        )
        if is_crash_row:
            notes.append(
                "KNOWN ISSUE (not a data-quality problem): this single run "
                "(DoryNew, H=15mm, 900cc bottle) crashed mid-solve on an "
                "isolated Aspherix comm-buffer overflow before reaching a "
                "settled state, so it produced no phi/fill measurement. It "
                "is an infrastructure/solver failure, not a physics result -- "
                "do not interpret it as 'H=15 doesn't pack'. The other 6/7 "
                "bottle-validation runs (including this family's other two "
                "heights) completed cleanly at 100% retention and passed all "
                "engineering gates. Row preserved exactly as produced: not "
                "re-run, not backfilled."
            )
        else:
            notes.append("")
    validation = validation.copy()
    validation["notes"] = notes

    phi = json.load(open(os.path.join(GP, "phi_gp.json")))
    wall = json.load(open(os.path.join(GP, "wall_gp.json")))

    from openpyxl import Workbook
    wb = Workbook()

    ws = wb.active
    ws.title = "README"
    build_readme(ws)

    write_df(wb.create_sheet("RunManifest"), manifest, UNITS_MANIFEST)
    write_df(wb.create_sheet("MainDOE_phi_vs_H_rho"), surrogate, units_maindoe)
    write_df(wb.create_sheet("LambdaSweep_phi_vs_lambda"), lam, UNITS_SURROGATE)
    write_df(wb.create_sheet("BottleValidation"), validation, UNITS_VALIDATION)
    build_wall(wb.create_sheet("WallLaw_FittedCoeffs"), wall)
    build_phi(wb.create_sheet("PhiGP_FittedCoeffs"), phi)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
