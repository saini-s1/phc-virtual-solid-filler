import openpyxl, warnings
from pathlib import Path
warnings.filterwarnings("ignore")
PATH = Path(__file__).resolve().parent.parent / "docs" / "nutrition-reference" / "Nutrition Calculator.template .xlsx"
wbF = openpyxl.load_workbook(PATH, data_only=False)  # formulas
wbV = openpyxl.load_workbook(PATH, data_only=True)   # cached values

def dump(name, only_text=False):
    ws = wbF[name]; wv = wbV[name]
    print(f"\n===== {name}  ({ws.dimensions}) =====")
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if hasattr(v, "text"): v = v.text
            if v is None: continue
            s = str(v).strip()
            if not s: continue
            val = wv[c.coordinate].value
            extra = f"   [={val}]" if (isinstance(v,str) and v.startswith("=") and val is not None) else ""
            print(f"  {c.coordinate}: {s}{extra}")

dump("Nutrition STONF")
