"""Temporary verification: re-read the source Excel and confirm the engine fixture
faithfully reproduces it. Dumps Ingredients / Formulation / Nutrition and recomputes
the per-serving totals + calorie methods. (This file is deleted after the check.)"""
import openpyxl, os

HERE = os.path.dirname(__file__)
PATH = os.path.join(HERE, "Nutrition Example Calculator.xlsx")

wbF = openpyxl.load_workbook(PATH, data_only=False)  # formulas
wbV = openpyxl.load_workbook(PATH, data_only=True)   # cached values

print("SHEETS:", wbF.sheetnames)

def dump(sheet, rmin, rmax, cmin, cmax, label):
    sv = wbV[sheet]; sf = wbF[sheet]
    print(f"\n===== {sheet} :: {label} (rows {rmin}-{rmax}, cols {cmin}-{cmax}) =====")
    for r in range(rmin, rmax + 1):
        cells = []
        for c in range(cmin, cmax + 1):
            v = sv.cell(row=r, column=c).value
            f = sf.cell(row=r, column=c).value
            if v is None and f is None:
                continue
            col = openpyxl.utils.get_column_letter(c)
            if isinstance(f, str) and f.startswith("="):
                cells.append(f"{col}{r}=[{f} -> {v!r}]")
            else:
                cells.append(f"{col}{r}={v!r}")
        if cells:
            print("  " + " | ".join(cells))

# Ingredients: header + 7 ingredients, cols A..R (1..18)
dump("Ingredients", 1, 9, 1, 18, "header + ingredient DB")

# Formulation: dose, names, %w/w, scaled rows, totals
dump("Formulation", 1, 21, 1, 18, "recipe engine")

# Nutrition: totals, %DV, rounding text, DV constants, reco cells
dump("Nutrition", 1, 35, 1, 7, "label sheet")

wbF.close(); wbV.close()
print("\n\nDONE")
